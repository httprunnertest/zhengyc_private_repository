import wx
import wx.grid
import os
import xlrd
import openpyxl as ol


class MyFrame(wx.Frame):
    def __init__(self, parent=None):
        super(MyFrame, self).__init__(parent, -1, "测试版", size=(1600, 900))

        # 初始化
        self.path = './data'
        self.current_file = ''
        self.current_sheet = ''
        self.rangeDot = []
        self.SetIcon(wx.Icon(name="./gui/flower.jpg", type=wx.BITMAP_TYPE_JPEG))
        self.current_pitch = ()
        self.rowNum = 0
        self.sheet_num = 0

        # 面板设置
        self.panel_left = wx.Panel(self, -1)
        self.panel_left_right = wx.Panel(self, -1)
        panel_top = wx.Panel(self, -1)
        panel_middle = wx.Panel(self, -1)
        self.panel_left.SetBackgroundColour('#FFFFFF')
        vbox = wx.BoxSizer(wx.VERTICAL)
        hbox = wx.BoxSizer(wx.HORIZONTAL)
        hbox.Add(self.panel_left, 1, wx.EXPAND)
        hbox.Add(self.panel_left_right, 1, wx.EXPAND)
        vbox.Add(panel_top, 1, wx.EXPAND | wx.ALIGN_LEFT)
        vbox.Add(panel_middle, 12, wx.EXPAND | wx.ALIGN_LEFT)
        hbox.Add(vbox, 6, wx.EXPAND)
        self.SetSizer(hbox)
        """
            这里是左边的两个选择器
        """

        # 表格设置
        gridbox = wx.BoxSizer(wx.VERTICAL)
        self.gridtable = wx.grid.Grid(panel_middle, -1, size=(400, 200))
        gridbox.Add(self.gridtable, 1, wx.EXPAND)
        panel_middle.SetSizer(gridbox)
        self.gridSetting()

        # 按钮设置
        self.addButton = wx.Button(panel_top, wx.ID_ANY, '新增', (0, 10), (100, 30))
        self.deleteButton = wx.Button(panel_top, wx.ID_ANY, '删除', (120, 10), (100, 30))
        self.saveButton = wx.Button(panel_top, wx.ID_ANY, '保存', (240, 10), (100, 30))
        self.rangeButton = wx.Button(panel_top, wx.ID_ANY, '合并单元格', (360, 10), (100, 30))
        self.testButton = wx.Button(panel_top, wx.ID_ANY, '开始测试', (480, 10), (100, 30))

        # 初始化
        self.bind_event()
        self.menu()
        self.select_file_and_sheet()
        self.sheet_list.Bind(wx.EVT_RIGHT_DOWN, self.right_down)

    def select_file_and_sheet(self):
        filename = [i[2] for i in os.walk(self.path)][0]
        font = wx.Font(18, wx.DECORATIVE, wx.NORMAL, wx.BOLD)
        self.file_list = wx.ListBox(parent=self.panel_left, id=-1, choices=filename, size=(230, 7000),
                                    style=wx.LB_SINGLE | wx.LB_SORT)
        self.file_list.SetFont(font)
        self.Bind(wx.EVT_LISTBOX, self.choose_file, self.file_list)
        #  根据文件名获取excel的sheet 名
        self.sheet_list = wx.ListBox(parent=self.panel_left_right, id=-1, choices=[], size=(230, 7000),
                                     style=wx.LB_SINGLE | wx.LB_SORT)
        self.sheet_list.SetFont(font)
        self.Bind(wx.EVT_LISTBOX, self.choose_sheet, self.sheet_list)

    def bind_event(self):
        self.Bind(wx.grid.EVT_GRID_RANGE_SELECT, self.OnRangeSelect)
        self.Bind(wx.grid.EVT_GRID_CELL_LEFT_CLICK, self.current_row, self.gridtable)
        self.Bind(wx.EVT_BUTTON, self.delete, self.deleteButton)
        self.Bind(wx.EVT_BUTTON, self.add, self.addButton)
        self.Bind(wx.EVT_BUTTON, self.save, self.saveButton)
        self.Bind(wx.EVT_BUTTON, self.range, self.rangeButton)
        self.Bind(wx.EVT_BUTTON, self.test, self.testButton)

    def menu(self):
        menuBar = wx.MenuBar()
        filemenu_help = wx.Menu()
        filemenu_file = wx.Menu()
        '''
        生成菜单类对象 wx.Menu()
        利用Append()方法直接添加菜单项
        参数1：id ID_ABOUT 是关于的标准id，推荐优先使用标准id
        参数2：字符串，菜单名
        参数3：选中菜单项时的提示信息
        '''
        menuItemAbout = filemenu_help.Append(wx.ID_ABOUT, "About", " Information about this program")
        '''
        Bind 绑定事件
        参数1：事件类型    wx.EVT_MENU 菜单被选择时的事件
        参数2：绑定的事件  self.OnAbout  自定义
        参数3：绑定的菜单项 menuItemAbout 在菜单添加的时候的返回值
        '''
        menuItemRegister = filemenu_help.Append(wx.ID_ANY, 'Register', 'Enter registration data')
        # AppendSeparator() 菜单项之前添加一个横线分隔
        filemenu_help.AppendSeparator()
        menuItemExit = filemenu_help.Append(wx.ID_EXIT, "Exit", "Terminate the program")
        self.Bind(wx.EVT_MENU, self.OnExit, menuItemExit)

        filemenu_file_item = wx.MenuItem(filemenu_file, wx.ID_SAVE, text='Save', helpString='Save the current word',
                                         kind=wx.ITEM_NORMAL)
        filemenu_file.Append(filemenu_file_item)
        menuItemNew = filemenu_file.Append(wx.ID_NEW, 'New', 'Create new class,interface,file or directory',
                                           wx.ITEM_NORMAL)
        filemenu_file.Append(wx.ID_OPEN, 'Open', 'Open a project or a file in editor', wx.ITEM_NORMAL)
        menuBar.Append(filemenu_file, 'File')
        menuBar.Append(filemenu_help, "Help")  # Adding the "filemenu" to the MenuBar
        self.SetMenuBar(menuBar)
        self.Bind(wx.EVT_MENU, self.New, menuItemNew)
        # 右键菜单
        self.right_menu = wx.Menu()
        self.right_menu.Append(-1, '新增sheet')
        self.right_menu.Append(-1, '删除sheet')

    def gridSetting(self):
        self.gridtable.CreateGrid(self.rowNum, 6)
        self.gridtable.SetColLabelValue(0, '操作名称')
        self.gridtable.SetColLabelValue(1, '操作步骤')
        self.gridtable.SetColLabelValue(2, '定位方式')
        self.gridtable.SetColLabelValue(3, '定位内容')
        self.gridtable.SetColLabelValue(4, '操作')
        self.gridtable.SetColLabelValue(5, '输入值/定位')
        self.gridtable.SetDefaultRowSize(40, 10)
        self.gridtable.SetDefaultColSize(170, 100)

    def current_row(self, e):
        self.current_pitch = (e.GetRow(), e.GetCol())
        e.Skip()

    def delete(self, e):
        try:
            self.gridtable.DeleteRows(pos=self.current_pitch[0])
            self.rowNum = self.rowNum - 1
        except IndexError:
            wx.MessageBox('请先选中值', '提示', wx.OK | wx.ICON_INFORMATION)

    def add(self, e):
        self.gridtable.AppendRows()
        self.rowNum = self.rowNum + 1

    def save(self, e):
        content = []
        for row in range(self.rowNum):
            cntlist = []
            for col in range(6):
                cntlist.append(self.gridtable.GetCellValue(row, col))
            content.append(cntlist)
        content.insert(0, [self.gridtable.GetColLabelValue(i) for i in range(6)])
        self.write_in_excel(f'{self.path}/{self.current_file}', self.current_sheet, content)

    def range(self, e):
        rows = self.rangeDot[1][0] - self.rangeDot[0][0] + 1
        cols = self.rangeDot[1][1] - self.rangeDot[0][1] + 1
        self.gridtable.SetCellSize(self.rangeDot[0][0], self.rangeDot[0][1], rows, cols)

    def test(self, e):
        # os.system('python run.py')
        collables = [self.gridtable.GetColLabelValue(i) for i in range(6)]
        # print(collables)

    def OnExit(self, e):
        self.Close(True)  # Close the frame.

    def New(self, e):
        dlg = wx.TextEntryDialog(None, "新增文件名:", "新增")
        if dlg.ShowModal() == wx.ID_OK:
            message = dlg.GetValue()
            file = ol.Workbook()
            envsheet = file.create_sheet('env')
            envsheet.cell(1, 1, '浏览器')
            envsheet.cell(1, 2, '坏境')
            del file['Sheet']
            file.save(f'{self.path}/{message}.xlsx')
            self.file_list.Append(f'{message}.xlsx')
        dlg.Destroy()

    def temp(self, e):
        """
            temp 方法用来往表中插入数据
            参数e,若为事件调用则，即wx._core.CommandEvent类,输入缓存的数据
                若为外部调用则为二维的数组类  list类
        """
        data = []
        if isinstance(e, list):
            data = e
            if len(data) > self.rowNum:
                for i in range(len(data) - self.rowNum - 1):
                    self.add('增加行')
            if len(data) < self.rowNum:
                for i in range(self.rowNum - len(data) + 1):
                    self.gridtable.DeleteRows()
                self.rowNum = len(data) - 1
        if data:
            data.pop(0)
        self.gridtable.ClearGrid()
        for i, j in enumerate(data):
            for k in range(len(data[0])):
                self.gridtable.SetCellValue(i, k, str(j[k]))

    # def getTempRow(self):
    #     with open('temp.csv', 'r', encoding='utf8')as init:
    #         readers = csv.reader(init)
    #         data = [row for row in readers]
    #         return data

    def OnRangeSelect(self, e):
        self.rangeDot = [e.GetTopLeftCoords(), e.GetBottomRightCoords()]
        e.Skip()

    def choose_file(self, e):
        listbox = e.GetEventObject().GetStringSelection()
        self.current_file = listbox
        self.sheet_list.Clear()
        self.sheet_num = len(self.get_excel_sheet_name(f'{self.path}/{listbox}'))
        data = sorted(self.get_excel_sheet_name(f'{self.path}/{listbox}'))
        self.sheet_list.Append(data)

    def choose_sheet(self, e):
        listbox = e.GetEventObject().GetStringSelection()
        self.current_sheet = listbox
        data = self.get_excel_sheet_data(f'{self.path}/{self.current_file}', listbox)
        print(data)
        self.updata_collabel(data[0])
        self.temp(data)

    def get_excel_sheet_data(self, path, sheetname):
        sheet_data = []
        file = xlrd.open_workbook(path)
        data = file.sheet_by_name(sheetname)
        for i in range(0, data.nrows):
            sheet_data.append(data.row_values(i))
        return sheet_data

    def get_excel_sheet_name(self, path):
        file = xlrd.open_workbook(path)
        return file.sheet_names()

    def updata_collabel(self, collables):
        for i in range(6):
            self.gridtable.SetColLabelValue(i, '')
        for index, lable in enumerate(collables):
            self.gridtable.SetColLabelValue(index, str(lable))

    def write_in_excel(self, path, sheetname, data):
        file = ol.load_workbook(path)
        del file[sheetname]
        sheet = file.create_sheet(sheetname)
        for row_index, i in enumerate(data):
            for col_index, j in enumerate(i):
                sheet.cell(row_index + 1, col_index + 1, j)
        file.save(path)

    def right_down(self, e):
        if 0 <= e.y <= 24 * self.sheet_num:
            self.sheet_list.PopupMenu(self.right_menu, pos=(e.x, e.y))
            dat = self.get_excel_sheet_name(f'{self.path}/{self.current_file}')
            data = sorted([i.lower() for i in dat])
            for index, k in enumerate(dat):
                if k.lower() == data[int(e.y / 24)]:
                    print(dat[index])


if __name__ == '__main__':
    app = wx.App()
    frame = MyFrame()
    frame.Show()
    app.MainLoop()
